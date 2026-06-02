from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.models import QuestionType
from app.parsers.common import sort_questions_by_number
from app.schemas import ParseResult, QuestionCreate


def parse_xlsx(path: Path) -> ParseResult:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ParseResult(title=path.stem, questions=[])

    header = [str(c or "").strip().lower() for c in rows[0]]
    col_map = {}
    aliases = {
        "stem": ["题干", "题目", "问题", "stem", "question"],
        "type": ["类型", "题型", "type"],
        "options": ["选项", "options"],
        "answer": ["答案", "正确答案", "answer"],
        "explanation": ["解析", "说明", "explanation"],
    }
    for key, names in aliases.items():
        for i, h in enumerate(header):
            if any(n in h for n in names):
                col_map[key] = i
                break

    questions: list[QuestionCreate] = []
    for idx, row in enumerate(rows[1:], start=0):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        cells = [str(c or "").strip() for c in row]

        def get_col(name: str, default: str = "") -> str:
            i = col_map.get(name)
            return cells[i] if i is not None and i < len(cells) else default

        stem = get_col("stem") or (cells[0] if cells else "")
        if not stem:
            continue
        answer = get_col("answer") or (cells[-1] if len(cells) > 1 else "")
        opts_raw = get_col("options")
        options = [o.strip() for o in opts_raw.split("|") if o.strip()] if opts_raw else None
        if not options and len(cells) >= 4:
            maybe_opts = [c for c in cells[1:-2] if c and not c.startswith("答案")]
            if len(maybe_opts) >= 2:
                options = maybe_opts

        type_raw = get_col("type").lower()
        type_map = {
            "单选": QuestionType.single,
            "单选题": QuestionType.single,
            "多选": QuestionType.multiple,
            "多选题": QuestionType.multiple,
            "判断": QuestionType.judge,
            "判断题": QuestionType.judge,
            "填空": QuestionType.fill,
            "填空题": QuestionType.fill,
            "简答": QuestionType.short,
            "简答题": QuestionType.short,
        }
        qtype = type_map.get(type_raw)
        if not qtype:
            from app.parsers.common import detect_type

            qtype = detect_type(stem, options or [], answer)

        questions.append(
            QuestionCreate(
                type=qtype,
                stem=stem,
                options=options,
                answer=answer,
                explanation=get_col("explanation") or None,
                order_index=idx,
            )
        )
    wb.close()
    return ParseResult(title=path.stem, questions=sort_questions_by_number(questions))
